import requests
import bs4
import openpyxl
from io import BytesIO


def load_workbook_from_url(url):
    response = requests.get(url)
    wb = openpyxl.load_workbook(BytesIO(response.content))
    return wb


def getLatestRestaurantReport():
    url = 'https://www.sanantonio.gov/Health/News/RestaurantReports'
    response = requests.get(url)
    soup = bs4.BeautifulSoup(response.text, 'html.parser')
    table = soup.select('#lt-229314082-2021')
    # Get last link
    last_link = table[0].select('a')[-1]['href']
    absolute_url = 'https://www.sanantonio.gov' + last_link
    # print(absolute_url)
    wb = load_workbook_from_url(absolute_url)
    ws = wb[wb.sheetnames[0]]
    return ws


sourceFile = getLatestRestaurantReport()

destinationFile = '2021.xlsx'
wb2 = openpyxl.load_workbook(destinationFile)
ws2 = wb2.active

mr = sourceFile.max_row
mc = sourceFile.max_column

for i in range(1, mr + 1):
    for j in range(1, mc + 1):
        # print(sourceFile.cell(row=i, column=j).value)
        ws2.cell(row=i, column=j).value = sourceFile.cell(
            row=i, column=j).value

wb2.save(destinationFile)
