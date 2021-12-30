import requests
import bs4
import openpyxl
from io import BytesIO
import datetime

# Save the current date and time so that we may prepend the update with timestamp later on.
now = datetime.datetime.now()

# This function allows us to load an xlsx file from a URL for use by openpyxl.


def load_workbook_from_url(url):
    response = requests.get(url)
    wb = openpyxl.load_workbook(BytesIO(response.content))
    return wb

# This function adds a period to a string if it is missing


def addPeriod(observation):
    if '.' not in observation:
        return observation + '.'
    else:
        return observation

# This function gets the latest restaurant report from the city of San Antonio's website..


def getLatestRestaurantReport():
    url = 'https://www.sanantonio.gov/Health/News/RestaurantReports'
    response = requests.get(url)
    soup = bs4.BeautifulSoup(response.text, 'html.parser')
    table = soup.select('#lt-229314082-2021')
    last_link = table[0].select('a')[-1]['href']
    absolute_url = 'https://www.sanantonio.gov' + last_link
    wb = load_workbook_from_url(absolute_url)
    ws = wb[wb.sheetnames[0]]
    return ws

# This function grabs all of the individual report links from the main report and returns them as a list.


def getInspections():
    inspectionList = []
    restaurantReport = getLatestRestaurantReport()
    nInspections = restaurantReport.max_row
    for inspection in range(2, nInspections + 1):
        if restaurantReport.cell(row=inspection, column=6).value < 90:
            inspectionList.append(restaurantReport.cell(
                row=inspection, column=7).hyperlink.target)
    return inspectionList

# This rat nest of a function grabs data from the individual inspection reports, including specific observations, and appends it to the markdown file.


def getInspectionDetails(f='', content=''):
    restaurantInfoList = []
    for inspection in getInspections():
        num = 0
        response = requests.get(inspection)
        soup = bs4.BeautifulSoup(response.text, 'html.parser')
        inspection_date = soup.find_all(
            'td')[3].get_text().strip()[5:]
        restaurant_name = soup.find_all(
            'td')[13].get_text().strip()[19:].title()
        repeat_violations = soup.find_all(
            'td')[15].find('strong').get_text().strip()[-1]
        score = soup.find_all(
            'td')[16].get_text().strip()
        address = soup.find_all(
            'td')[17].get_text().strip()[18:]
        observationTable = soup.select(
            '#container > #main > .padL')[2]

        observationDetails = observationTable.find_all('td')
        observations = []
        for observation in observationDetails[4:]:
            num += 1
            if num % 2 != 0:
                observations.append(observation.get_text().strip())

        observationList = []
        for observation in range(0, len(observations)):
            ob = observations[observation].split('. ')
            # print(ob)
            for obs in range(len(ob)):
                # print(ob[obs])
                if ob[obs].lower().startswith(' **** observed:'):
                    partitions = ob[obs].lower().partition('observed:')
                    result = partitions[1] + partitions[2]
                    observationList.append(addPeriod(result.capitalize()))
                elif ob[obs].lower().startswith('**** observed:'):
                    partitions = ob[obs].lower().partition('observed:')
                    result = partitions[1] + partitions[2]
                    observationList.append(addPeriod(result.capitalize()))
                elif ob[obs].lower().startswith('observed'):
                    extracted = ob[obs].split('\t')
                    observationList.append(
                        addPeriod(extracted[0].capitalize()))
                elif 'observed' in ob[obs].lower():
                    partitions = ob[obs].lower().partition('observed')
                    result = partitions[1] + partitions[2]
                    observationList.append(addPeriod(result.capitalize()))
                elif ob[obs].lower().startswith('observed '):
                    observationList.append(addPeriod(ob[obs].capitalize()))

        if len(observationList) == 0:
            observationList = 'None'
        else:
            observationList = '\n* '.join(observationList)

        print('Restaurant Name: ' + restaurant_name)
        restaurantInfoList.append('## {}\n**Inspection date:** {}\n\n**Score:** {}\n\n**Address:** {}\n\n**Repeat violations:** {}\n\n[Full report]({})\n\n**Notable observations:**\n* {}\n'.format(
            restaurant_name, inspection_date, score, address, repeat_violations, inspection, observationList))
        print('--------')
    f.write('Updated: {}\n'.format(str(now)) +
            '\n'.join(restaurantInfoList) + '\n***\n' + content)


with open('story.md', "r+") as f:
    content = f.read()
    f.seek(0)
    getInspectionDetails(f, content)
