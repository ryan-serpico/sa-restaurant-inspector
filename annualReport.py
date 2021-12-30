import requests
import bs4
import openpyxl
from io import BytesIO
import pandas as pd

# This fuction allows us to read an excel file from a url using openpyxl.


def load_workbook_from_url(url):
    response = requests.get(url)
    wb = openpyxl.load_workbook(BytesIO(response.content))
    return wb

# This function returns a list of all the restaurant report URLs.


def getRestaurantReports():
    url = 'https://www.sanantonio.gov/Health/News/RestaurantReports'
    response = requests.get(url)
    soup = bs4.BeautifulSoup(response.text, 'html.parser')
    table = soup.select('#lt-229314082-2021')
    # Find all links
    linkList = []
    links = table[0].select('a')
    for link in links:
        absolute_url = 'https://www.sanantonio.gov' + link['href']
        linkList.append(absolute_url)
    return linkList


# This function concatenates all the restaurant report data into one file.
def combineData():
    counter = 0
    columns = ['ESTABLISHMENT NAME', 'ESTABLISHMENT ADDRESS',
               'INSPECTION DATE',	'SECTOR',	'DISTRICT',	'TOTAL SCORE',	'LINK', 'Link']

    all_data = pd.DataFrame(columns=columns)
    for f in getRestaurantReports():

        counter += 1
        print(counter)
        df = pd.read_excel(f)

        wb = load_workbook_from_url(f)
        ws = wb[wb.sheetnames[0]]
        links = []
        for i in range(2, ws.max_row + 1):
            links.append(ws.cell(row=i, column=7).hyperlink.target)
        df['Link'] = links
        all_data = pd.concat([all_data, df])
    all_data.drop('LINK', axis=1, inplace=True)
    all_data = all_data.sort_values(by=['TOTAL SCORE'])
    all_data.to_excel('combined.xlsx', index=False)


combineData()
